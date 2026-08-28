import io
import re
from datetime import date, datetime, timedelta

from pypdf import PdfReader
from openpyxl import load_workbook


# --------------------------------------------------
# 기본 설정
# --------------------------------------------------

STOPWORDS = {
    "알려줘", "알려", "뭐야", "무엇", "뭐", "어떤",
    "관련", "대한", "있는", "있어", "있나요",
    "업무", "일정", "내용",
    "이번", "이번주", "주에",
    "해줘", "줘",
}


DEADLINE_WORDS = {
    "제출",
    "마감",
    "완료",
    "점검",
    "회의",
    "보고",
    "작성",
    "검토",
    "확인",
    "납기"
}


def clean_text(text):
    return re.sub(
        r"\s+",
        " ",
        str(text or "")
    ).strip()


# --------------------------------------------------
# 날짜 처리
# --------------------------------------------------

def parse_date(text, today=None):

    today = today or date.today()

    text = clean_text(text)

    patterns = [

        # 2026-09-03
        # 2026.09.03
        # 2026년 9월 3일
        r"(?P<y>20\d{2})[.\-/년]\s*"
        r"(?P<m>\d{1,2})[.\-/월]\s*"
        r"(?P<d>\d{1,2})일?",

        # 9월 3일
        r"(?P<m>\d{1,2})월\s*"
        r"(?P<d>\d{1,2})일",

        # 9/3
        # 9-3
        # 9.3
        r"(?P<m>\d{1,2})[./-]"
        r"(?P<d>\d{1,2})",
    ]

    for pattern in patterns:

        match = re.search(
            pattern,
            text
        )

        if not match:
            continue

        groups = match.groupdict()

        year = (
            int(groups["y"])
            if groups.get("y")
            else today.year
        )

        month = int(
            groups["m"]
        )

        day = int(
            groups["d"]
        )

        try:

            result = date(
                year,
                month,
                day
            )

            # 연도가 없는 일정인데
            # 현재 날짜보다 너무 오래 전이면
            # 다음 해 일정으로 간주
            if (
                not groups.get("y")
                and
                result < today - timedelta(days=90)
            ):

                result = date(
                    today.year + 1,
                    month,
                    day
                )

            return result

        except ValueError:
            pass

    return None


def format_excel_value(value):

    if value is None:
        return ""

    if isinstance(
        value,
        (datetime, date)
    ):

        return (
            f"{value.month}/"
            f"{value.day}"
        )

    return clean_text(
        value
    )


# --------------------------------------------------
# Excel 열 이름 찾기
# --------------------------------------------------

def find_column(
    headers,
    candidates
):

    for candidate in candidates:

        for index, header in enumerate(headers):

            if candidate == header:
                return index

    return None


# --------------------------------------------------
# Excel 읽기
# --------------------------------------------------

def load_excel(uploaded_file):

    chunks = []

    workbook = load_workbook(
        io.BytesIO(
            uploaded_file.getvalue()
        ),
        data_only=True
    )

    for sheet in workbook.worksheets:

        rows = list(
            sheet.iter_rows(
                values_only=True
            )
        )

        if not rows:
            continue

        # ------------------------------
        # 1행 = 제목으로 사용
        # ------------------------------

        headers = [
            clean_text(value)
            for value in rows[0]
        ]

        date_col = find_column(
            headers,
            [
                "일자",
                "날짜",
                "일정"
            ]
        )

        task_col = find_column(
            headers,
            [
                "업무",
                "업무명",
                "내용",
                "일정 내용"
            ]
        )

        owner_col = find_column(
            headers,
            [
                "담당",
                "담당자",
                "담당 부서"
            ]
        )

        status_col = find_column(
            headers,
            [
                "상태",
                "진행상태",
                "진행 상태"
            ]
        )

        note_col = find_column(
            headers,
            [
                "비고",
                "메모",
                "특이사항"
            ]
        )

        # ------------------------------
        # 실제 데이터
        # ------------------------------

        for row_number, row in enumerate(
            rows[1:],
            start=2
        ):

            values = [
                format_excel_value(value)
                for value in row
            ]

            if not any(values):
                continue

            def get_value(index):

                if index is None:
                    return ""

                if index >= len(values):
                    return ""

                return values[index]

            record = {

                "일자":
                    get_value(date_col),

                "업무":
                    get_value(task_col),

                "담당":
                    get_value(owner_col),

                "상태":
                    get_value(status_col),

                "비고":
                    get_value(note_col)
            }

            # 기존 검색과의 호환을 위한 전체 텍스트
            text = " | ".join(
                [
                    value
                    for value in record.values()
                    if value
                ]
            )

            chunks.append({

                "text": text,

                "source":
                    uploaded_file.name,

                "location":
                    f"{sheet.title} 시트 "
                    f"{row_number}행",

                "type":
                    "xlsx",

                # ⭐ 핵심
                # 각 열을 따로 저장
                "record":
                    record,

                "date":
                    parse_date(
                        record["일자"]
                    )
            })

    return chunks


# --------------------------------------------------
# PDF
# --------------------------------------------------

def load_pdf(uploaded_file):

    chunks = []

    reader = PdfReader(
        io.BytesIO(
            uploaded_file.getvalue()
        )
    )

    for page_number, page in enumerate(
        reader.pages,
        start=1
    ):

        text = (
            page.extract_text()
            or ""
        )

        if not text.strip():
            continue

        chunks.append({

            "text": text,

            "source":
                uploaded_file.name,

            "location":
                f"p.{page_number}",

            "type":
                "pdf"
        })

    return chunks


# --------------------------------------------------
# TXT
# --------------------------------------------------

def load_txt(uploaded_file):

    chunks = []

    text = (
        uploaded_file
        .getvalue()
        .decode(
            "utf-8-sig",
            errors="ignore"
        )
    )

    for line_number, line in enumerate(
        text.splitlines(),
        start=1
    ):

        if not line.strip():
            continue

        chunks.append({

            "text":
                line.strip(),

            "source":
                uploaded_file.name,

            "location":
                f"{line_number}행",

            "type":
                "txt"
        })

    return chunks


# --------------------------------------------------
# 지식 DB 생성
# --------------------------------------------------

def build_knowledge_base(
    uploaded_files
):

    knowledge = []

    for uploaded_file in uploaded_files:

        filename = (
            uploaded_file.name.lower()
        )

        if filename.endswith(
            ".xlsx"
        ):

            knowledge.extend(
                load_excel(
                    uploaded_file
                )
            )

        elif filename.endswith(
            ".pdf"
        ):

            knowledge.extend(
                load_pdf(
                    uploaded_file
                )
            )

        elif filename.endswith(
            ".txt"
        ):

            knowledge.extend(
                load_txt(
                    uploaded_file
                )
            )

    return knowledge


# --------------------------------------------------
# 질문 의도 분석
# --------------------------------------------------

def detect_intent(question):

    q = question.replace(
        " ",
        ""
    )

    if any(
        word in q
        for word in [
            "담당",
            "담당자",
            "누가",
            "누구"
        ]
    ):
        return "담당"

    if any(
        word in q
        for word in [
            "상태",
            "진행상태",
            "진행중",
            "완료됐",
            "완료된"
        ]
    ):
        return "상태"

    if any(
        word in q
        for word in [
            "비고",
            "특이사항",
            "주의사항",
            "메모"
        ]
    ):
        return "비고"

    if any(
        word in q
        for word in [
            "언제",
            "날짜",
            "일자",
            "마감일"
        ]
    ):
        return "일자"

    return "전체"


# --------------------------------------------------
# 질문 키워드
# --------------------------------------------------

def get_question_keywords(
    question
):

    words = re.findall(
        r"[가-힣A-Za-z0-9]+",
        question.lower()
    )

    keywords = []

    for word in words:

        if word in STOPWORDS:
            continue

        # 날짜 숫자는 따로 처리
        if re.fullmatch(
            r"\d+",
            word
        ):
            continue

        if len(word) >= 2:
            keywords.append(
                word
            )

    return keywords


# --------------------------------------------------
# 검색
# --------------------------------------------------

def search_knowledge(
    question,
    knowledge,
    today=None,
    top_k=5
):

    today = today or date.today()

    question_date = parse_date(
        question,
        today=today
    )

    keywords = get_question_keywords(
        question
    )

    wants_week = (
        "이번 주" in question
        or
        "이번주" in question
    )

    wants_priority = any(
        word in question
        for word in [
            "먼저",
            "우선",
            "급한",
            "최우선",
            "가장 먼저"
        ]
    )

    week_end = (
        today
        +
        timedelta(
            days=6 - today.weekday()
        )
    )

    results = []

    for item in knowledge:

        score = 0

        # ==================================
        # Excel 구조화 데이터
        # ==================================

        if (
            item.get("type")
            == "xlsx"
            and
            item.get("record")
        ):

            record = item["record"]

            item_date = (
                item.get("date")
                or
                parse_date(
                    record.get(
                        "일자",
                        ""
                    ),
                    today=today
                )
            )

            task = record.get(
                "업무",
                ""
            )

            owner = record.get(
                "담당",
                ""
            )

            status = record.get(
                "상태",
                ""
            )

            note = record.get(
                "비고",
                ""
            )

            # --------------------------
            # 날짜 질문은 최우선
            # --------------------------

            if question_date:

                if (
                    item_date
                    == question_date
                ):
                    score += 150

                else:
                    # 다른 날짜는 거의 제외
                    score -= 100

            # --------------------------
            # 키워드 검색
            # --------------------------

            for keyword in keywords:

                if keyword in task.lower():
                    score += 30

                if keyword in owner.lower():
                    score += 25

                if keyword in status.lower():
                    score += 25

                if keyword in note.lower():
                    score += 20

            # --------------------------
            # 진행 상태 질문
            # --------------------------

            for state_word in [
                "진행중",
                "완료",
                "예정",
                "미완료"
            ]:

                if (
                    state_word in question
                    and
                    state_word in status
                ):

                    score += 80

            # --------------------------
            # 최우선 검색
            # --------------------------

            if (
                "최우선" in question
                and
                "최우선" in note
            ):
                score += 100

            # --------------------------
            # 이번 주
            # --------------------------

            if (
                wants_week
                and
                item_date
            ):

                if (
                    today
                    <= item_date
                    <= week_end
                ):

                    score += 80

            # --------------------------
            # 먼저 해야 할 업무
            # --------------------------

            if (
                wants_priority
                and
                item_date
            ):

                days_left = (
                    item_date
                    -
                    today
                ).days

                if (
                    0
                    <= days_left
                    <= 30
                ):

                    score += (
                        30 - days_left
                    )

                if "최우선" in note:
                    score += 60

            if score > 0:

                copied = item.copy()

                copied[
                    "score"
                ] = score

                copied[
                    "date"
                ] = item_date

                results.append(
                    copied
                )

        # ==================================
        # PDF / TXT
        # ==================================

        else:

            text = clean_text(
                item.get(
                    "text",
                    ""
                )
            )

            for keyword in keywords:

                if (
                    keyword.lower()
                    in text.lower()
                ):
                    score += 10

            item_date = parse_date(
                text,
                today=today
            )

            if question_date:

                if (
                    item_date
                    == question_date
                ):
                    score += 100

            if score > 0:

                copied = item.copy()

                copied[
                    "score"
                ] = score

                copied[
                    "date"
                ] = item_date

                results.append(
                    copied
                )

    results.sort(
        key=lambda x: (
            -x["score"],
            x.get("date")
            or date.max
        )
    )

    return results[
        :top_k
    ]


# --------------------------------------------------
# 날짜 표시
# --------------------------------------------------

def date_label(
    value
):

    if not value:
        return ""

    return (
        f"{value.month}월 "
        f"{value.day}일"
    )


# --------------------------------------------------
# Excel 답변 생성
# --------------------------------------------------

def make_excel_answer(
    question,
    results
):

    intent = detect_intent(
        question
    )

    if not results:

        return (
            "조건에 맞는 업무를 "
            "찾지 못했습니다."
        )

    # 점수 차이가 큰 경우
    # 가장 정확한 결과만 사용
    best_score = results[0][
        "score"
    ]

    strong_results = [
        result
        for result in results
        if result["score"]
        >= best_score - 10
    ]

    # 최대 5개
    strong_results = (
        strong_results[:5]
    )

    # ------------------------------
    # 하나의 업무
    # ------------------------------

    if len(
        strong_results
    ) == 1:

        result = (
            strong_results[0]
        )

        record = result[
            "record"
        ]

        d = date_label(
            result.get(
                "date"
            )
        )

        task = record.get(
            "업무",
            ""
        )

        owner = record.get(
            "담당",
            ""
        )

        status = record.get(
            "상태",
            ""
        )

        note = record.get(
            "비고",
            ""
        )

        if intent == "담당":

            return (
                f"{d} {task} 업무의 "
                f"담당은 {owner}입니다."
            )

        if intent == "상태":

            return (
                f"{d} {task} 업무의 "
                f"현재 상태는 "
                f"{status}입니다."
            )

        if intent == "비고":

            if note:

                return (
                    f"{d} {task} 업무의 "
                    f"비고는 "
                    f"'{note}'입니다."
                )

            return (
                f"{d} {task} 업무에는 "
                "별도로 등록된 "
                "비고가 없습니다."
            )

        if intent == "일자":

            return (
                f"{task} 업무의 "
                f"일자는 {d}입니다."
            )

        # 기본 전체 답변
        answer = (
            f"{d}에는 "
            f"{task} 업무가 있습니다."
        )

        if owner:
            answer += (
                f" 담당은 "
                f"{owner}입니다."
            )

        if status:
            answer += (
                f" 현재 상태는 "
                f"{status}입니다."
            )

        if note:
            answer += (
                f" 비고는 "
                f"'{note}'입니다."
            )

        return answer

    # ------------------------------
    # 여러 업무
    # ------------------------------

    lines = [
        "조건에 맞는 업무를 "
        f"{len(strong_results)}건 "
        "찾았습니다."
    ]

    for result in strong_results:

        record = result[
            "record"
        ]

        d = date_label(
            result.get(
                "date"
            )
        )

        task = record.get(
            "업무",
            ""
        )

        owner = record.get(
            "담당",
            ""
        )

        status = record.get(
            "상태",
            ""
        )

        line = (
            f"- {d}: {task}"
        )

        extras = []

        if owner:
            extras.append(
                f"담당 {owner}"
            )

        if status:
            extras.append(
                f"상태 {status}"
            )

        if extras:

            line += (
                " ("
                +
                ", ".join(
                    extras
                )
                +
                ")"
            )

        lines.append(
            line
        )

    return "\n".join(
        lines
    )


# --------------------------------------------------
# 최종 답변 생성
# --------------------------------------------------

def make_answer(
    question,
    results,
    today=None
):

    today = today or date.today()

    if not results:

        return (
            "관련된 근거 문서를 "
            "찾지 못했습니다. "
            "질문을 조금 더 "
            "구체적으로 입력해주세요."
        )

    # Excel 결과가 최상위면
    # 구조화된 답변 사용
    excel_results = [
        result
        for result in results
        if (
            result.get("type")
            == "xlsx"
            and
            result.get("record")
        )
    ]

    if excel_results:

        return make_excel_answer(
            question,
            excel_results
        )

    # PDF / TXT 기본 답변
    best = results[0]

    text = clean_text(
        best.get(
            "text",
            ""
        )
    )

    found_date = best.get(
        "date"
    )

    if found_date:

        return (
            f"{date_label(found_date)} "
            f"관련 문서에서 다음 내용을 "
            f"확인했습니다: {text[:200]}"
        )

    return (
        "관련 문서에서 다음 내용을 "
        f"확인했습니다: {text[:200]}"
    )